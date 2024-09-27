from django.shortcuts import render, redirect
from .models import Empleados, Asistencias, Salario, Lote
from proyectos.models import Proyectos
from contabilidad.models import GastosManoObra
from .forms import EmpleadosForm, AsistenciaExcelForm, NominasExcelForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from datetime import datetime, date
from django.db.models import Q
from decimal import Decimal
from django.db.models import Sum
from django.db.models.functions import Coalesce

import openpyxl

# Create your views here.

class NominasListView(LoginRequiredMixin,ListView):
    model = Salario
    template_name = "empleados/nominas_list.html"
    context_object_name = "nominas"

    def get_queryset(self):
        slug = self.kwargs.get('slug')
        lote_id = self.kwargs.get('lote')
        proyecto = Proyectos.objects.get(slug=slug)
        lote = Lote.objects.get(id=lote_id)
        return Salario.objects.filter(proyecto=proyecto,lote=lote.id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        slug = self.kwargs.get('slug')
        proyecto = Proyectos.objects.get(slug=slug)
        context['lote'] = self.kwargs.get('lote')
        context['proyecto'] = proyecto
        return context 

class AsistenciasListView(LoginRequiredMixin,ListView):
    model = Asistencias
    template_name = "empleados/asistencias.html"
    context_object_name = "asistencias"

    def get_queryset(self):
        slug = self.kwargs.get('slug')
        proyecto = Proyectos.objects.get(slug=slug)
        return Asistencias.objects.filter(proyecto=proyecto)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        slug = self.kwargs.get('slug')
        proyecto = Proyectos.objects.get(slug=slug)

        context['proyecto'] = proyecto
        return context 

class EmpleadosListView(LoginRequiredMixin,ListView):
    model = Empleados
    template_name = "empleados/empleados.html"
    context_object_name = "empleados"

    def get_queryset(self):
        # Obtener la queryset inicial
        queryset = Empleados.objects.all()

        # Obtener los parámetros de filtro de la URL o de un formulario GET
        empleado = self.request.GET.get('empleado')
        rfc = self.request.GET.get('rfc')
        imss = self.request.GET.get('imss')
        infonavit = self.request.GET.get('infonavit')

        # Aplicar los filtros si los valores no son nulos
        if empleado:
            queryset = queryset.filter(empleado=empleado)
        
        if rfc:
            queryset = queryset.filter(rfc__icontains=rfc)
        
        if imss:
            queryset = queryset.filter(imss__icontains=imss)

        if infonavit:
            queryset = queryset.filter(infonavit__icontains=infonavit)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener el slug del proyecto desde la URL
        # slug = self.kwargs.get('slug')
        # Si necesitas obtener el proyecto con base en el slug:
        # proyecto = Proyectos.objects.get(slug=slug)

        # Pasar los filtros actuales para mantener el estado del formulario
        context['empleado'] = self.request.GET.get('empleado', '')
        context['rfc'] = self.request.GET.get('rfc', '')
        context['imss'] = self.request.GET.get('imss', '')
        context['infonavit'] = self.request.GET.get('infonavit', '')
        # context['project'] = proyecto

        return context


@login_required
def registro_empleados(request):

    empleados = Empleados.objects.all()

    if request.method == "POST":
        empleados_form = EmpleadosForm(request.POST)

        if empleados_form.is_valid():
            empleados_form.save()
            return redirect('empleados:empleados')
    else:
        empleados_form = EmpleadosForm()

    return render(request,'empleados/registrar_empleados.html',{'empleados_form':empleados_form,'empleados':empleados})


def registro_asistencias(request):
    if request.method == 'POST':
        form = AsistenciaExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            workbook = openpyxl.load_workbook(archivo_excel)

            # Cargar las dos hojas: "Asistencias" y "Horas Extras"
            hoja_asistencias = workbook['Asistencias']
            hoja_horas_extras = workbook['Horas Extras']

            # Extraer el proyecto desde la celda A1 de la hoja de asistencias
            clave_proyecto = hoja_asistencias['B1'].value
            print(f"Proyecto ID: {clave_proyecto}")  # Depuración

            try:
                proyecto = Proyectos.objects.get(clave_proyecto=clave_proyecto, estatus=True)
                print(f"Proyecto encontrado: {proyecto.proyecto}")  # Depuración
            except Proyectos.DoesNotExist:
                print("El proyecto no existe o no está activo.")  # Depuración
                return render(request, 'error.html', {'mensaje': 'El proyecto no existe o no está activo.'})

            
            # Diccionario para almacenar los datos de asistencias y horas extras
            diccionario_asistencias = {
                "empleado": [],
                "asistencias": [],
                "horas_extras": [],
                "fecha": []
            }

            # Procesar asistencias desde la hoja "Asistencias"
            for fila in hoja_asistencias.iter_rows(min_row=3, max_col=hoja_asistencias.max_column, values_only=True):
                nombre_empleado = fila[0]
                try:
                    empleado = Empleados.objects.get(rfc=nombre_empleado)
                except Empleados.DoesNotExist:
                    continue  # Si no existe el empleado, omitir

                for idx, valor in enumerate(fila[1:], start=1):
                    if valor is not None and valor == 1:  # Solo registrar asistencias si el valor es 1
                        fecha_asistencia = convertir_fecha(hoja_asistencias.cell(row=2, column=idx + 1).value)
                        
                        # Filtrar fechas inválidas
                        if not fecha_asistencia or fecha_asistencia > datetime.now():
                            continue
                        
                        # Verificar duplicados
                        asistencia_existente = Asistencias.objects.filter(
                            Q(empleado=empleado) & Q(fecha=fecha_asistencia)
                        ).exists()

                        if not asistencia_existente:
                            # Añadir al diccionario
                            diccionario_asistencias["empleado"].append(empleado)
                            diccionario_asistencias["asistencias"].append(True)
                            diccionario_asistencias["horas_extras"].append(Decimal('0.00'))  # Inicialmente sin horas extras
                            diccionario_asistencias["fecha"].append(fecha_asistencia)

            # Procesar horas extras desde la hoja "Horas Extras"
            for fila in hoja_horas_extras.iter_rows(min_row=3, max_col=hoja_horas_extras.max_column, values_only=True):
                nombre_empleado = fila[0]
                try:
                    empleado = Empleados.objects.get(rfc=nombre_empleado)
                except Empleados.DoesNotExist:
                    continue  # Si no existe el empleado, omitir

                for idx, valor in enumerate(fila[1:], start=1):
                    if valor is not None and valor > 0:  # Solo registrar horas extras si hay un valor mayor a 0
                        fecha_extras = convertir_fecha(hoja_horas_extras.cell(row=2, column=idx + 1).value)

                        # Filtrar fechas inválidas
                        if not fecha_extras or fecha_extras > datetime.now():
                            continue

                        # Verificar si ya existe una asistencia registrada en el diccionario
                        for i, (emp, fecha) in enumerate(zip(diccionario_asistencias["empleado"], diccionario_asistencias["fecha"])):
                            if emp == empleado and fecha == fecha_extras:
                                # Actualizar las horas extras
                                diccionario_asistencias["horas_extras"][i] = valor
                                break

            # Crear una lista de objetos Asistencias
            asistencias_a_crear = []

            for i in range(len(diccionario_asistencias["empleado"])):
                asistencia = Asistencias(
                    empleado=diccionario_asistencias["empleado"][i],
                    proyecto=proyecto,  # Debes definir el proyecto
                    asistencias=diccionario_asistencias["asistencias"][i],
                    horas_extras=diccionario_asistencias["horas_extras"][i],
                    fecha=diccionario_asistencias["fecha"][i]
                )
                asistencias_a_crear.append(asistencia)

            # Guardar todos los objetos en la base de datos de una sola vez
            Asistencias.objects.bulk_create(asistencias_a_crear)

            # Redirigir a la vista de asistencias del proyecto
            return redirect('empleados:asistencias', slug=proyecto.slug)

    else:
        form = AsistenciaExcelForm()

    return render(request, 'empleados/registrar_asistencias.html', {'asistencias_form': form})

# Función auxiliar para convertir la fecha del formato Excel a datetime
def convertir_fecha(fecha_excel):
    if isinstance(fecha_excel, date):
        return fecha_excel
    elif isinstance(fecha_excel, datetime):
        return fecha_excel.date()
    elif isinstance(fecha_excel, str):
        try:
            return datetime.strptime(fecha_excel, '%Y-%m-%d').date()
        except ValueError:
            print(f"Error: formato de fecha no válido en cadena: {fecha_excel}")
            return None
    elif isinstance(fecha_excel, (int, float)):
        try:
            return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(fecha_excel) - 2).date()
        except Exception as e:
            print(f"Error: no se pudo convertir fecha numérica: {fecha_excel}. Error: {e}")
            return None
    else:
        print(f"Error: formato de fecha no reconocido: {fecha_excel}")
        return None

def registro_nominas(request,slug):
    if request.method == 'POST':
        form = NominasExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            workbook = openpyxl.load_workbook(archivo_excel)

            # Cargar las dos hojas: "Nóminas"
            hoja_nominas = workbook['Nominas']

            # Extraer el proyecto desde la celda B1 de la hoja de nóminas
            # clave_proyecto = hoja_nominas['B1'].value
            proyecto = Proyectos.objects.get(slug=slug)
            clave_proyecto = proyecto.clave_proyecto
            print(f"Proyecto ID: {clave_proyecto}")  # Depuración

            try:
                proyecto = Proyectos.objects.get(clave_proyecto=clave_proyecto, estatus=True)
                print(f"Proyecto encontrado: {proyecto.proyecto}")  # Depuración
            except Proyectos.DoesNotExist:
                print("El proyecto no existe o no está activo.")  # Depuración
                return render(request, 'error.html', {'mensaje': 'El proyecto no existe o no está activo.'})
            
            # Diccionario para almacenar los datos de nóminas y horas extras
            diccionario_nominas = {
                "empleado": [],
                "salario": [],
                "infonavit": [],
                "imss": [],
                "isr": [],
                "horas_extras": []
            }

            # Procesar asistencias desde la hoja "Nóminas"
            for fila in hoja_nominas.iter_rows(min_row=3, max_col=hoja_nominas.max_column, values_only=True):
                nombre_empleado = fila[0]
                try:
                    empleado = Empleados.objects.get(rfc=nombre_empleado)
                    # Asumimos que los datos están en las columnas correctas como se ve en la plantilla
                    diccionario_nominas["empleado"].append(empleado)
                    diccionario_nominas["salario"].append(Decimal(fila[1]))
                    diccionario_nominas["infonavit"].append(Decimal(fila[2]))
                    diccionario_nominas["imss"].append(Decimal(fila[3]))
                    diccionario_nominas["isr"].append(Decimal(fila[4]))
                    diccionario_nominas["horas_extras"].append(Decimal(fila[5]))
                except Empleados.DoesNotExist:
                    continue  # Si no existe el empleado, omitir
            
            # Crear una lista de objetos Nominas para usar en bulk_create
            nuevo_lote = Lote.obtener_nuevo_lote(proyecto=proyecto)
            nominas_a_crear = []
            for i in range(len(diccionario_nominas["empleado"])):
                nomina = Salario(
                    empleado=diccionario_nominas["empleado"][i],
                    proyecto=proyecto,
                    salario=diccionario_nominas["salario"][i],
                    infonavit=diccionario_nominas["infonavit"][i],
                    imss=diccionario_nominas["imss"][i],
                    isr=diccionario_nominas["isr"][i],
                    horas_extras=diccionario_nominas["horas_extras"][i],
                    lote=nuevo_lote.id 
                    ) 
                
                nominas_a_crear.append(nomina)

            Salario.objects.bulk_create(nominas_a_crear)

            registro_gastos_mano_obra(proyecto,nuevo_lote.id)
            
            # Redirigir a la vista de asistencias del proyecto
            return redirect('empleados:nominas', slug=proyecto.slug, lote=nuevo_lote.id)

    else:
        form = NominasExcelForm()

    return render(request, 'empleados/registro_nominas.html',{'nominas_form':form})


def registro_gastos_mano_obra(proyecto,lote):
    nominas_result = Salario.objects.filter(lote=lote,proyecto=proyecto).aggregate(
        total_nominas=Coalesce(Sum('salario'),Decimal('0.00'))
    )
    nominas=nominas_result['total_nominas']
    infonavit_result = Salario.objects.filter(lote=lote,proyecto=proyecto).aggregate(
        total_infonavit=Coalesce(Sum('infonavit'),Decimal('0.00'))
    )
    infonavit=infonavit_result['total_infonavit']
    imss_result = Salario.objects.filter(lote=lote,proyecto=proyecto).aggregate(
        total_imss=Coalesce(Sum('imss'),Decimal('0.00'))
    )
    imss=imss_result['total_imss']
    isr_result = Salario.objects.filter(lote=lote,proyecto=proyecto).aggregate(
        total_isr=Coalesce(Sum('isr'),Decimal('0.00'))
    )
    isr=isr_result['total_isr']
    horas_extras_result = Salario.objects.filter(lote=lote,proyecto=proyecto).aggregate(
        total_horas_extras=Coalesce(Sum('horas_extras'),Decimal('0.00'))
    )
    horas_extras=horas_extras_result['total_horas_extras']
    
    isn = Decimal('0.03') * nominas
    total_monto = nominas + infonavit + imss + isr + horas_extras + isn

    GastosManoObra.objects.create(
        proyecto=proyecto,
        nomina=nominas,
        infonavit=infonavit,
        imss=imss,
        isn=isn,
        isr=isr,
        horas_extras=horas_extras,
        monto=total_monto,
    )